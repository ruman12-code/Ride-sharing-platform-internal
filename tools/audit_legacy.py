#!/usr/bin/env python3
"""
Legacy artefact audit harness.

Reads the macro-enabled workbook that the carpooling process ran on for five
months and emits the raw material for docs/LEGACY_AUDIT.md:

  out/vba/*.bas          one file per extracted VBA module, verbatim
  out/sheets/*.csv       one CSV per worksheet, cells as stored (no coercion)
  out/audit_rows.json    per-row field capture + defect flags + disposition
  out/audit_summary.json counts, per-column type profile, defect tallies

Design rules
  - Read only. The workbook is evidence; we never write to it.
  - No coercion on read. A cell holding the text "06:60 am" is recorded as the
    string "06:60 am", not silently parsed, dropped, or NaN'd. Coercion is what
    destroyed the provenance of the legacy date column in the first place.
  - Every defect is attached to a row, and every row gets a disposition, so the
    audit is reviewable line by line rather than in aggregate.

Usage
  python3 tools/audit_legacy.py legacy/Ride_sharing_platformFinal29012024.xlsm
"""

from __future__ import annotations

import csv
import datetime as dt
import json
import re
import sys
import zipfile
from pathlib import Path
from typing import Any

OUT = Path("out")

# ---------------------------------------------------------------------------
# Defect detection
#
# Each rule is (code, description, predicate). Predicates take the raw cell
# value and return True when the defect is present. Kept declarative so the
# audit document can be generated from this table rather than hand-written and
# drifting out of sync with the code that produced it.
# ---------------------------------------------------------------------------

TIME_RE = re.compile(r"^\s*(\d{1,2})\s*:\s*(\d{2})\s*(am|pm)?\s*$", re.I)
INT_RE = re.compile(r"^\s*\d+\s*$")


def is_blank(v: Any) -> bool:
    return v is None or (isinstance(v, str) and v.strip() == "")


def impossible_time(v: Any) -> bool:
    """Text that looks like a time but cannot be one, e.g. '06:60 am'."""
    if not isinstance(v, str):
        return False
    m = TIME_RE.match(v)
    if not m:
        return False
    hour, minute = int(m.group(1)), int(m.group(2))
    meridiem = m.group(3)
    if minute > 59:
        return True
    if meridiem and not (1 <= hour <= 12):
        return True
    if not meridiem and hour > 23:
        return True
    return False


def unparseable_time(v: Any) -> bool:
    """Neither a real time value nor text matching any time shape at all."""
    if isinstance(v, (dt.time, dt.datetime)):
        return False
    if is_blank(v):
        return False
    return isinstance(v, str) and TIME_RE.match(v) is None


def non_numeric_seats(v: Any) -> bool:
    """Seat counts that are not integers, e.g. the literal string 'plenty'."""
    if is_blank(v):
        return False
    if isinstance(v, (int, float)) and float(v).is_integer():
        return False
    return not (isinstance(v, str) and INT_RE.match(v))


def text_date(v: Any) -> bool:
    """A date stored as text rather than a real date serial.

    The legacy column mixes dd/mm/yyyy text with genuine Excel dates. Once
    mixed, the two are indistinguishable without the original author's intent:
    03/04/2024 is either 3 April or 4 March and nothing in the file resolves it.
    Flagged so every affected row is quarantined rather than guessed at.
    """
    return isinstance(v, str) and not is_blank(v)


# Names that are not plausible colleague names: comic-book characters, a
# cricketer, a singer, an internet meme. Listed explicitly rather than inferred,
# because "is this a real person" is a judgement a reviewer must be able to
# challenge row by row, not a heuristic hidden in a regex.
FICTIONAL_PERSONAS = {
    "spiderman", "batman", "superman", "voyager-1", "mr. t",
    "frank sinatra", "dwaine bravo", "podda nodir majhi",
}


def fictional_persona(v: Any) -> bool:
    if not isinstance(v, str):
        return False
    # Strip a parenthesised phone number before comparing: one row carries
    # "Dwaine Bravo (01722334455)" in the name field.
    name = re.sub(r"\s*\(.*?\)", "", v).strip().lower()
    return name in FICTIONAL_PERSONAS


def phone_in_name(v: Any) -> bool:
    """A phone number typed into the free-text name field. PII leak."""
    return isinstance(v, str) and re.search(r"\d{9,}", v) is not None


DEFECT_RULES = [
    ("BLANK_ROUTE", "Origin or destination empty", None),
    ("IMPOSSIBLE_TIME", "Time-shaped text that is not a valid time", impossible_time),
    ("UNPARSEABLE_TIME", "Departure time neither a time value nor time-shaped", unparseable_time),
    ("NON_NUMERIC_SEATS", "Seat count is not an integer", non_numeric_seats),
    ("TEXT_DATE", "Date stored as text; dd/mm vs mm/dd now ambiguous", text_date),
    ("DUPLICATE_POSTING", "Identical route+date+time as an earlier row", None),
    ("REPEAT_TRIP", "Same user re-entered the same route on another date", None),
    ("PAST_DATE", "Departure date precedes the row's own entry timestamp", None),
    ("FICTIONAL_PERSONA", "Poster name is not a plausible colleague", fictional_persona),
    ("PHONE_IN_NAME", "Phone number typed into the free-text name field", phone_in_name),
]


# ---------------------------------------------------------------------------
# Column identification
#
# The header row is read from the file rather than assumed. This map is a set of
# candidate substrings per logical field; whichever header matches wins. If a
# logical field finds no header, that is itself an audit finding (recorded in
# summary.unmapped_fields) rather than a crash.
# ---------------------------------------------------------------------------

FIELD_HINTS = {
    "entered_at":   ["timestamp", "entry", "submitted", "date of entry"],
    "user":         ["name", "employee", "staff", "posted by", "driver"],
    "email":        ["email", "mail"],
    "phone":        ["phone", "mobile", "contact", "cell"],
    "origin":       ["starting place", "origin", "from", "place", "pick", "start"],
    "destination":  ["to", "destination", "drop", "end"],
    "via":          ["via", "route", "through", "passing"],
    "date":         ["date"],
    "time":         ["time", "departure"],
    "seats":        ["seat", "capacity", "passenger", "space"],
    "vehicle":      ["vehicle", "car", "model"],
    "notes":        ["note", "comment", "remark"],
}


def map_columns(headers: list[str]) -> dict[str, int]:
    """Match logical fields to physical column indices.

    Scoring, highest first:
      3  hint matches the whole header    ("Destination" -> destination)
      2  hint matches a whole word in it  ("Starting Date" -> date, via "date")
      1  hint appears only as a substring ("Starting Date" -> origin, via "start")
    then by hint length.

    The tiers exist because a naive substring match is actively wrong on this
    workbook: "Starting Date" contains "start", so an origin hint steals the
    date column and the date field silently reads as unmapped. Requiring a word
    boundary lets the specific match (date) beat the incidental one (start).
    """
    candidates: list[tuple[int, int, str, int]] = []  # (tier, hintlen, field, idx)
    for field, hints in FIELD_HINTS.items():
        for idx, h in enumerate(headers):
            if not h:
                continue
            low = str(h).strip().lower()
            words = set(re.findall(r"[a-z]+", low))
            for hint in hints:
                if low == hint:
                    tier = 3
                elif hint in words or all(w in words for w in hint.split()):
                    tier = 2
                elif hint in low:
                    tier = 1
                else:
                    continue
                candidates.append((tier, len(hint), field, idx))

    candidates.sort(key=lambda c: (-c[0], -c[1]))
    mapping: dict[str, int] = {}
    taken: set[int] = set()
    for _tier, _len, field, idx in candidates:
        if field in mapping or idx in taken:
            continue
        mapping[field] = idx
        taken.add(idx)
    return mapping


# ---------------------------------------------------------------------------
# VBA extraction
# ---------------------------------------------------------------------------

def extract_vba(path: Path) -> dict[str, str]:
    """Pull every VBA module out of the workbook, verbatim."""
    try:
        from oletools.olevba import VBA_Parser
    except ImportError:
        sys.exit("oletools is required: pip install oletools")

    modules: dict[str, str] = {}
    parser = VBA_Parser(str(path))
    if not parser.detect_vba_macros():
        print("  ! no VBA macros detected in workbook", file=sys.stderr)
        return modules
    for _, _, vba_filename, vba_code in parser.extract_macros():
        name = Path(vba_filename).name or f"module_{len(modules)}"
        # A workbook can carry two modules of the same name in different
        # streams; keep both rather than letting one silently overwrite the
        # other, since divergent copies of one handler is a defect we expect.
        if name in modules:
            name = f"{name}.dup{len(modules)}"
        modules[name] = vba_code
    parser.close()
    return modules


def analyse_vba(modules: dict[str, str]) -> list[dict[str, Any]]:
    """Static findings over the extracted VBA.

    Deliberately shallow and textual. The goal is not to build a VBA compiler,
    it is to produce a defensible list of what is broken, each finding pointing
    at a module and a line so a reviewer can confirm it against the source.
    """
    findings: list[dict[str, Any]] = []

    for name, code in modules.items():
        lines = code.splitlines()
        proc_stack: list[tuple[int, str]] = []

        for i, line in enumerate(lines, start=1):
            stripped = line.strip()
            low = stripped.lower()

            # Nested procedure: a Sub/Function opened while another is still
            # open. VBA does not permit this and the module will not compile.
            if re.match(r"^(public |private |friend )?(sub|function)\s+\w+", low):
                if proc_stack:
                    findings.append({
                        "module": name, "line": i, "code": "VBA_NESTED_PROC",
                        "detail": f"{stripped!r} opens while "
                                  f"{proc_stack[-1][1]!r} (line {proc_stack[-1][0]}) "
                                  f"is still open; module cannot compile",
                    })
                proc_stack.append((i, stripped))
            elif re.match(r"^end (sub|function)\b", low):
                if proc_stack:
                    proc_stack.pop()

            # Sheet addressed by display name in quotes. Breaks the moment
            # anyone renames the tab; the code-name form does not.
            for m in re.finditer(r'Sheets\(\s*"([^"]+)"\s*\)|Worksheets\(\s*"([^"]+)"\s*\)', stripped):
                findings.append({
                    "module": name, "line": i, "code": "VBA_SHEET_BY_STRING",
                    "detail": f"sheet addressed as {(m.group(1) or m.group(2))!r} "
                              f"(display name) rather than by code name",
                })

            # The append pattern that silently overwrites under concurrency.
            if re.search(r"lastrow\s*\+\s*1", low) or re.search(r"\.rows\.count.*end\(xlup\)", low):
                findings.append({
                    "module": name, "line": i, "code": "VBA_LASTROW_APPEND",
                    "detail": "computes next free row then writes; two users "
                              "submitting concurrently write the same row and "
                              "one posting is lost with no error",
                })

            if "instr(" in low:
                findings.append({
                    "module": name, "line": i, "code": "VBA_SUBSTRING_SEARCH",
                    "detail": "InStr substring match used for location search; "
                              "matches on origin text only, no destination or "
                              "corridor awareness",
                })

            if low.startswith("on error resume next"):
                findings.append({
                    "module": name, "line": i, "code": "VBA_ERROR_SWALLOWED",
                    "detail": "On Error Resume Next suppresses all failures, "
                              "including failed writes",
                })

        if proc_stack:
            for ln, txt in proc_stack:
                findings.append({
                    "module": name, "line": ln, "code": "VBA_UNCLOSED_PROC",
                    "detail": f"{txt!r} never closed by End Sub/End Function",
                })

    # Divergent duplicate handlers: same procedure name, different bodies.
    bodies: dict[str, list[tuple[str, str]]] = {}
    for name, code in modules.items():
        for m in re.finditer(
            r"^(?:public |private )?(?:sub|function)\s+(\w+)(.*?)^end (?:sub|function)",
            code, re.I | re.M | re.S,
        ):
            bodies.setdefault(m.group(1).lower(), []).append((name, m.group(2)))
    for proc, occurrences in bodies.items():
        if len(occurrences) > 1:
            distinct = {re.sub(r"\s+", " ", b).strip() for _, b in occurrences}
            findings.append({
                "module": ", ".join(n for n, _ in occurrences),
                "line": 0,
                "code": "VBA_DUPLICATE_PROC" if len(distinct) == 1 else "VBA_DIVERGENT_PROC",
                "detail": f"{proc!r} defined {len(occurrences)} times; "
                          f"{'identical' if len(distinct) == 1 else 'bodies DIFFER'}",
            })

    return findings


# ---------------------------------------------------------------------------
# Sheet reading
# ---------------------------------------------------------------------------

def read_sheets(path: Path) -> dict[str, list[list[Any]]]:
    """Every sheet, every cell, as stored. No type coercion, no header skipping."""
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True, read_only=True)
    sheets: dict[str, list[list[Any]]] = {}
    for ws in wb.worksheets:
        sheets[ws.title] = [list(row) for row in ws.iter_rows(values_only=True)]
    wb.close()
    return sheets


def profile_row(row: list[Any], cols: dict[str, int]) -> dict[str, Any]:
    """Pull the logical fields out of one physical row."""
    def get(field: str) -> Any:
        idx = cols.get(field)
        if idx is None or idx >= len(row):
            return None
        return row[idx]
    return {f: get(f) for f in FIELD_HINTS}


def jsonable(v: Any) -> Any:
    if isinstance(v, (dt.datetime, dt.date, dt.time)):
        return v.isoformat()
    if v is None or isinstance(v, (str, int, float, bool)):
        return v
    return str(v)


def audit_rows(sheet_name: str, rows: list[list[Any]]) -> tuple[list[dict], dict]:
    if not rows:
        return [], {"error": "sheet is empty"}

    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    cols = map_columns(headers)
    unmapped = [f for f in FIELD_HINTS if f not in cols]

    seen_signature: dict[tuple, int] = {}
    seen_user_route: dict[tuple, list[int]] = {}
    out_rows: list[dict] = []

    for n, raw in enumerate(rows[1:], start=2):
        if all(is_blank(c) for c in raw):
            continue
        fields = profile_row(raw, cols)
        defects: list[str] = []

        if is_blank(fields["origin"]) or is_blank(fields["destination"]):
            defects.append("BLANK_ROUTE")
        if impossible_time(fields["time"]):
            defects.append("IMPOSSIBLE_TIME")
        if unparseable_time(fields["time"]):
            defects.append("UNPARSEABLE_TIME")
        if non_numeric_seats(fields["seats"]):
            defects.append("NON_NUMERIC_SEATS")
        if text_date(fields["date"]):
            defects.append("TEXT_DATE")
        if fictional_persona(fields["user"]):
            defects.append("FICTIONAL_PERSONA")
        if phone_in_name(fields["user"]):
            defects.append("PHONE_IN_NAME")

        sig = (
            str(fields["user"]).strip().lower(),
            str(fields["origin"]).strip().lower(),
            str(fields["destination"]).strip().lower(),
            str(fields["date"]).strip() if fields["date"] else "",
            str(fields["time"]).strip() if fields["time"] else "",
        )
        if sig in seen_signature:
            defects.append("DUPLICATE_POSTING")
        else:
            seen_signature[sig] = n

        route_key = (sig[0], sig[1], sig[2])
        seen_user_route.setdefault(route_key, []).append(n)

        # Disposition. Conservative by design: anything we cannot place on the
        # timeline unambiguously is repaired under review, never auto-migrated,
        # because a wrong date silently corrupts the liquidity baseline.
        if "FICTIONAL_PERSONA" in defects:
            disposition, why = "discard-as-test", "poster is not a real colleague"
        elif "BLANK_ROUTE" in defects:
            disposition, why = "discard-as-test", "no route; nothing to migrate"
        elif {"IMPOSSIBLE_TIME", "UNPARSEABLE_TIME", "NON_NUMERIC_SEATS", "TEXT_DATE"} & set(defects):
            disposition, why = "repair", "typed field holds untyped value; needs human adjudication"
        elif "DUPLICATE_POSTING" in defects:
            disposition, why = "discard-as-test", "exact duplicate of an earlier row"
        else:
            disposition, why = "migrate", "all required fields present and well-typed"

        out_rows.append({
            "sheet": sheet_name,
            "row": n,
            "fields": {k: jsonable(v) for k, v in fields.items()},
            "defects": defects,
            "disposition": disposition,
            "rationale": why,
        })

    # Second pass: repeat trips are only visible once every row is read.
    for route_key, row_numbers in seen_user_route.items():
        if len(row_numbers) > 1 and route_key[0] not in ("", "none"):
            for rn in row_numbers:
                for r in out_rows:
                    if r["row"] == rn and "DUPLICATE_POSTING" not in r["defects"]:
                        r["defects"].append("REPEAT_TRIP")

    defect_counts: dict[str, int] = {}
    for r in out_rows:
        for d in r["defects"]:
            defect_counts[d] = defect_counts.get(d, 0) + 1

    disposition_counts: dict[str, int] = {}
    for r in out_rows:
        disposition_counts[r["disposition"]] = disposition_counts.get(r["disposition"], 0) + 1

    summary = {
        "sheet": sheet_name,
        "headers": headers,
        "column_mapping": cols,
        "unmapped_fields": unmapped,
        "data_rows": len(out_rows),
        "defect_counts": defect_counts,
        "disposition_counts": disposition_counts,
        "repeat_routes": {
            " | ".join(k): v for k, v in seen_user_route.items() if len(v) > 1
        },
    }
    return out_rows, summary


def main() -> int:
    if len(sys.argv) != 2:
        print(__doc__)
        return 2
    path = Path(sys.argv[1])
    if not path.exists():
        print(f"ERROR: legacy workbook not found at {path}", file=sys.stderr)
        print("Place the .xlsm at that path and re-run. The audit cannot be "
              "written from anything other than the file itself.", file=sys.stderr)
        return 1

    (OUT / "vba").mkdir(parents=True, exist_ok=True)
    (OUT / "sheets").mkdir(parents=True, exist_ok=True)

    print(f"reading {path} ({path.stat().st_size:,} bytes)")

    print("extracting VBA ...")
    modules = extract_vba(path)
    for name, code in modules.items():
        (OUT / "vba" / name).write_text(code, encoding="utf-8")
        print(f"  {name}: {len(code.splitlines())} lines")
    vba_findings = analyse_vba(modules)
    print(f"  {len(vba_findings)} static findings")

    print("reading sheets ...")
    sheets = read_sheets(path)
    all_rows: list[dict] = []
    summaries: list[dict] = []
    for name, rows in sheets.items():
        safe = re.sub(r"[^A-Za-z0-9_-]", "_", name)
        with (OUT / "sheets" / f"{safe}.csv").open("w", newline="", encoding="utf-8") as fh:
            w = csv.writer(fh)
            for r in rows:
                w.writerow([jsonable(c) for c in r])
        rows_out, summary = audit_rows(name, rows)
        all_rows.extend(rows_out)
        summaries.append(summary)
        print(f"  {name!r}: {len(rows)} raw rows, {summary.get('data_rows', 0)} data rows")

    (OUT / "audit_rows.json").write_text(json.dumps(all_rows, indent=2), encoding="utf-8")
    (OUT / "audit_summary.json").write_text(json.dumps({
        "source_file": str(path),
        "source_bytes": path.stat().st_size,
        "generated_at": dt.datetime.now(dt.timezone.utc).isoformat(),
        "vba_modules": {n: len(c.splitlines()) for n, c in modules.items()},
        "vba_findings": vba_findings,
        "sheets": summaries,
    }, indent=2), encoding="utf-8")

    print(f"\nwrote out/audit_rows.json ({len(all_rows)} rows) and out/audit_summary.json")
    print("next: python3 tools/liquidity_baseline.py")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
